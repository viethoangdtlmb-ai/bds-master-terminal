# -*- coding: utf-8 -*-
"""Phan tich chuyen sau: Correlation, Lead-Lag, Momentum Divergence, Regime"""
import sys, io, csv, math
from collections import defaultdict
if sys.stdout.encoding != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
import openpyxl

# ============================================================
# LOAD DATA
# ============================================================
# CC data
cc_ts = defaultdict(list)  # {district: [prices by month]}
with open(r"d:\1. BDS\AI-Assistant\App_kham_benh_tai_san\gia_chung_cu_ha_noi_22_quan_60thang.csv", encoding="utf-8-sig") as f:
    for row in csv.DictReader(f):
        d = row["Khu_vuc"].strip()
        p = row["Gia_TB (tr/m2)"].strip()
        if p and p != "N/A":
            cc_ts[d].append(float(p))

# NR data
name_map_vn = {}  # vn -> ascii
wb = openpyxl.load_workbook(r"d:\1. BDS\AI-Assistant\06_Kien_Thuc_Du_An\Kham_benh_Ha_Noi\Gia_Nha_Rieng_2021_2026.xlsx", data_only=True)
ws = wb.active
districts_nr = [ws.cell(1, c).value for c in range(3, ws.max_column + 1)]
nr_ts = {}
for i, d in enumerate(districts_nr):
    vals = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, i + 3).value
        vals.append(float(v) if isinstance(v, (int, float)) and v > 0 else None)
    nr_ts[d] = vals

# Name mapping
nm = {
    "Ba Dinh": "Ba Đình", "Bac Tu Liem": "Bắc Từ Liêm", "Cau Giay": "Cầu Giấy",
    "Dan Phuong": "Đan Phượng", "Dong Anh": "Đông Anh", "Dong Da": "Đống Đa",
    "Gia Lam": "Gia Lâm", "Ha Dong": "Hà Đông", "Hai Ba Trung": "Hai Bà Trưng",
    "Hoai Duc": "Hoài Đức", "Hoan Kiem": "Hoàn Kiếm", "Hoang Mai": "Hoàng Mai",
    "Long Bien": "Long Biên", "Nam Tu Liem": "Nam Từ Liêm", "Tay Ho": "Tây Hồ",
    "Thanh Tri": "Thanh Trì", "Thanh Xuan": "Thanh Xuân", "Van Giang (HY)": "Văn giang",
}

# Rental data
rent = {}
with open(r"d:\1. BDS\AI-Assistant\06_Kien_Thuc_Du_An\Kham_benh_Ha_Noi\rental_yield_ha_noi.csv", encoding="utf-8-sig") as f:
    for row in csv.DictReader(f):
        n = row["Quận/Huyện"].strip()
        y = row["Rental Yield (%)"].strip()
        rent[n] = float(y) if y != "N/A" else 0

def returns(vals):
    """Tinh monthly returns tu price series"""
    r = []
    for i in range(1, len(vals)):
        if vals[i] and vals[i-1] and vals[i-1] > 0:
            r.append((vals[i] - vals[i-1]) / vals[i-1] * 100)
        else:
            r.append(0)
    return r

def correlation(a, b):
    """Pearson correlation"""
    n = min(len(a), len(b))
    if n < 10: return 0
    a, b = a[:n], b[:n]
    ma = sum(a)/n
    mb = sum(b)/n
    sa = math.sqrt(sum((x-ma)**2 for x in a)/n)
    sb = math.sqrt(sum((x-mb)**2 for x in b)/n)
    if sa == 0 or sb == 0: return 0
    return sum((a[i]-ma)*(b[i]-mb) for i in range(n)) / (n * sa * sb)

print("=" * 80)
print("  PHAN TICH CHUYEN SAU: CC × NR × THUE — HA NOI 2021-2026")
print("=" * 80)

# ============================================================
# 1. TUONG QUAN GIA CC vs NR (cung quan)
# ============================================================
print("\n📊 1. TUONG QUAN GIA CC vs NR (cung quan, 60 thang)")
print("-" * 80)
print(f"  {'Quan':<18} {'Corr CC-NR':>10} {'Nhan dinh'}")

corrs = []
for cc_name, cc_vals in sorted(cc_ts.items()):
    nr_name = nm.get(cc_name)
    if nr_name and nr_name in nr_ts:
        nr_vals = [v for v in nr_ts[nr_name] if v]
        n = min(len(cc_vals), len(nr_vals))
        if n >= 20:
            c = correlation(cc_vals[:n], nr_vals[:n])
            label = "🔗 Rat chat" if c > 0.95 else ("🔗 Chat" if c > 0.85 else ("🟡 Trung binh" if c > 0.7 else "🔴 Yeu"))
            corrs.append((cc_name, c, label))
            print(f"  {cc_name:<18} {c:>9.3f}   {label}")

avg_corr = sum(c for _, c, _ in corrs) / len(corrs) if corrs else 0
print(f"\n  TB toan thi truong: {avg_corr:.3f}")

# ============================================================
# 2. LEAD-LAG: CC dan NR hay NR dan CC?
# ============================================================
print("\n\n📊 2. LEAD-LAG: AI DAN AI? (CC dan NR bao nhieu thang?)")
print("-" * 80)

for cc_name, cc_vals in sorted(cc_ts.items()):
    nr_name = nm.get(cc_name)
    if nr_name and nr_name in nr_ts:
        nr_vals = [v for v in nr_ts[nr_name] if v is not None]
        cc_ret = returns(cc_vals)
        nr_ret = returns(nr_vals)
        n = min(len(cc_ret), len(nr_ret))
        if n < 20: continue
        
        best_lag = 0
        best_corr = correlation(cc_ret[:n], nr_ret[:n])
        
        for lag in range(1, 7):
            # CC leads NR by 'lag' months
            if n > lag:
                c = correlation(cc_ret[:n-lag], nr_ret[lag:n])
                if c > best_corr:
                    best_corr = c
                    best_lag = lag
            # NR leads CC
            if n > lag:
                c = correlation(nr_ret[:n-lag], cc_ret[lag:n])
                if c > best_corr:
                    best_corr = c
                    best_lag = -lag
        
        leader = f"CC dan {best_lag}T" if best_lag > 0 else (f"NR dan {-best_lag}T" if best_lag < 0 else "Dong thoi")
        print(f"  {cc_name:<18} {leader:<15} (corr={best_corr:.3f})")

# ============================================================
# 3. MOMENTUM DIVERGENCE (CC vs NR acceleration gap)
# ============================================================
print("\n\n📊 3. PHAN KY MOMENTUM: CC vs NR (12T gan)")
print("-" * 80)
print(f"  {'Quan':<18} {'CC 12T':>7} {'NR 12T':>7} {'Gap':>7} {'Tin hieu'}")

for cc_name, cc_vals in sorted(cc_ts.items()):
    nr_name = nm.get(cc_name)
    if nr_name and nr_name in nr_ts:
        nr_vals = [v for v in nr_ts[nr_name] if v is not None]
        if len(cc_vals) >= 13 and len(nr_vals) >= 13:
            cc_12m = ((cc_vals[-1] - cc_vals[-13]) / cc_vals[-13]) * 100
            nr_12m = ((nr_vals[-1] - nr_vals[-13]) / nr_vals[-13]) * 100
            gap = cc_12m - nr_12m
            if gap > 15: sig = "🔴 CC qua nong"
            elif gap > 5: sig = "⚠️ CC nhanh hon"
            elif gap > -5: sig = "✅ Can bang"
            elif gap > -15: sig = "🟡 NR nhanh hon"
            else: sig = "🏠 NR dan manh"
            print(f"  {cc_name:<18} {cc_12m:>+6.1f}% {nr_12m:>+6.1f}% {gap:>+6.1f}% {sig}")

# ============================================================
# 4. SHARPE RATIO: CC vs NR
# ============================================================
print("\n\n📊 4. SHARPE RATIO: CC vs NR (risk-adjusted)")
print("-" * 80)
print(f"  {'Quan':<18} {'CC Sharpe':>9} {'NR Sharpe':>9} {'Ai tot hon'}")

for cc_name, cc_vals in sorted(cc_ts.items()):
    nr_name = nm.get(cc_name)
    if nr_name and nr_name in nr_ts:
        nr_vals = [v for v in nr_ts[nr_name] if v is not None]
        cc_ret = returns(cc_vals)
        nr_ret = returns(nr_vals)
        
        if len(cc_ret) < 20 or len(nr_ret) < 20: continue
        
        cc_avg = sum(cc_ret)/len(cc_ret)
        cc_std = math.sqrt(sum((r-cc_avg)**2 for r in cc_ret)/len(cc_ret))
        cc_sharpe = cc_avg/cc_std if cc_std > 0 else 0
        
        nr_avg = sum(nr_ret)/len(nr_ret)
        nr_std = math.sqrt(sum((r-nr_avg)**2 for r in nr_ret)/len(nr_ret))
        nr_sharpe = nr_avg/nr_std if nr_std > 0 else 0
        
        better = "CC" if cc_sharpe > nr_sharpe else "NR"
        print(f"  {cc_name:<18} {cc_sharpe:>8.3f} {nr_sharpe:>9.3f}   {better}")

# ============================================================
# 5. COMPOSITE RANKING 3 CHIEU (recalc with real Sharpe)
# ============================================================
print("\n\n📊 5. XHANG TONG HOP (Growth CC + Growth NR + Yield + Sharpe CC)")
print("-" * 80)
print(f"  {'#':>2} {'Quan':<18} {'Diem':>5} {'CC G':>6} {'NR G':>6} {'Yield':>5} {'Sharpe':>6}")

ranking = []
for cc_name, cc_vals in cc_ts.items():
    nr_name = nm.get(cc_name)
    if not nr_name or nr_name not in nr_ts: continue
    nr_vals = [v for v in nr_ts[nr_name] if v is not None]
    if len(cc_vals) < 20 or len(nr_vals) < 20: continue
    
    cc_g = ((cc_vals[-1]-cc_vals[0])/cc_vals[0])*100
    nr_g = ((nr_vals[-1]-nr_vals[0])/nr_vals[0])*100
    yld = rent.get(cc_name, 0)
    
    cc_ret = returns(cc_vals)
    cc_avg = sum(cc_ret)/len(cc_ret)
    cc_std = math.sqrt(sum((r-cc_avg)**2 for r in cc_ret)/len(cc_ret))
    sharpe = cc_avg/cc_std if cc_std > 0 else 0
    
    # Score: 25% CC growth + 25% NR growth + 25% Yield + 25% Sharpe
    s_cc = min(cc_g / 400 * 25, 25)
    s_nr = min(nr_g / 300 * 25, 25)
    s_y = min(yld / 4 * 25, 25)
    s_sh = min(sharpe / 0.25 * 25, 25)
    total = s_cc + s_nr + s_y + s_sh
    
    ranking.append((cc_name, total, cc_g, nr_g, yld, sharpe))

ranking.sort(key=lambda x: x[1], reverse=True)
for i, (n, t, cg, ng, y, sh) in enumerate(ranking, 1):
    medal = "🥇" if i == 1 else ("🥈" if i == 2 else ("🥉" if i == 3 else f"{i:>2}"))
    print(f"  {medal} {n:<18} {t:>4.1f} {cg:>+5.0f}% {ng:>+5.0f}% {y:>4.1f}% {sh:>5.3f}")

print("\n" + "=" * 80)
