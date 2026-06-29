# -*- coding: utf-8 -*-
"""
crawl_lich_su_gia_cc.py
========================
Crawl lịch sử giá chung cư (tr/m²) theo THÁNG trong 5 năm.
Batdongsan.com.vn sử dụng kiến trúc Next.js nên dữ liệu được lưu ngầm 
(Hydrated state) tại thẻ <script id="__NEXT_DATA__">.

Cài đặt:
  pip install curl_cffi playwright openpyxl
  playwright install chromium

Chạy:
  python crawl_lich_su_gia_cc.py
"""

import sys, io, json, re, time, random
from datetime import datetime, date
from pathlib import Path

# Fix Windows console encoding
if sys.stdout.encoding != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

try:
    from playwright.sync_api import sync_playwright
    HAS_PLAYWRIGHT = True
except ImportError:
    HAS_PLAYWRIGHT = False
    print("❌ Playwright chưa cài. Chạy: pip install playwright && playwright install chromium")
    sys.exit(1)

try:
    import openpyxl
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

# ============================================================
# CONFIG
# ============================================================

OUTPUT_DIR = Path(__file__).parent

DISTRICTS = {
    "Nam Từ Liêm":   "nam-tu-liem",
    "Cầu Giấy":      "cau-giay",
    "Hà Đông":       "ha-dong",
    "Hoàng Mai":     "hoang-mai",
    "Thanh Xuân":    "thanh-xuan",
    "Tây Hồ":        "tay-ho",
    "Bắc Từ Liêm":  "bac-tu-liem",
    "Long Biên":     "long-bien",
    "Đống Đa":       "dong-da",
    "Hai Bà Trưng":  "hai-ba-trung",
    "Ba Đình":       "ba-dinh",
    "Hoàn Kiếm":     "hoan-kiem",
    "Gia Lâm":       "gia-lam",
    "Hoài Đức":      "hoai-duc",
    "Thanh Trì":     "thanh-tri",
    "Đông Anh":      "dong-anh",
    "Đan Phượng":    "dan-phuong",
    "Sóc Sơn":       "soc-son",
    "Mê Linh":       "me-linh",
    "Thạch Thất":    "thach-that",
    "Văn Giang (HY)": "van-giang",
    "Từ Sơn (BN)":   "tu-son",
}

CC_URL = "https://batdongsan.com.vn/ban-can-ho-chung-cu-{slug}"

# ============================================================
# EXTRACT DATA
# ============================================================

def safe_extract_prices(html: str) -> list[dict]:
    """Lấy dữ liệu giá qua biểu thức chính quy trực tiếp từ HTML để né parse JSON phức tạp của NEXT_DATA."""
    prices = []
    
    # Cách 1: Tìm dạng month-value array trực tiếp từ NextJS state stringified
    # Pattern nhắm tới: {"month":"2023-01","value":45.2}
    pts = re.findall(r'{"month"\s*:\s*"([2]\d{3}-\d{2})"\s*,\s*"value"\s*:\s*([\d.]+)}', html)
    if pts:
        for m, v in pts:
            prices.append({"thang": m, "gia_trm2": float(v)})
            
    # Cách 2: Tìm label và data arrays truyền thống của Highcharts/ChartJs
    # => labels: ["2020-01", ...], data: [45.2, ...]
    if not prices:
        m_labels = re.search(r'"labels"\s*:\s*(\[[^\]]+\])', html)
        m_data = re.search(r'"data"\s*:\s*(\[[^\]]+\])', html)
        if m_labels and m_data:
            try:
                lbs = json.loads(m_labels.group(1))
                dts = json.loads(m_data.group(1))
                if lbs and str(lbs[0]).startswith("20") and len(lbs) == len(dts):
                    for l, d in zip(lbs, dts):
                        if d is not None:
                            prices.append({"thang": str(l), "gia_trm2": float(d)})
            except: pass
            
    # Lọc duplicates bằng dict map
    unique_prices = {}
    for p in prices:
        unique_prices[p["thang"]] = p["gia_trm2"]
        
    # Chuyển về mảng và sort
    result = [{"thang": k, "gia_trm2": v} for k, v in unique_prices.items()]
    result.sort(key=lambda x: x["thang"])
    
    return result

# ============================================================
# CHROME BYPASS via Playwright
# ============================================================

def crawl_district_with_playwright(slug: str, name: str, playwright_ctx) -> list[dict]:
    url = CC_URL.format(slug=slug)
    print(f"\n  📍 {name} → {url}")

    try:
        page = playwright_ctx.new_page()
        page.route("**/*.{png,jpg,jpeg,gif,svg,woff,woff2,ttf,ico,css}", lambda route: route.abort())
        
        page.goto(url, wait_until="domcontentloaded", timeout=45000)
        
        # Đợi Cloudflare và Lazy components
        page.wait_for_timeout(5000)
        page.evaluate("window.scrollTo(0, document.body.scrollHeight / 3)")
        page.wait_for_timeout(2000)
        page.evaluate("window.scrollTo(0, document.body.scrollHeight / 2)")
        page.wait_for_timeout(2000)

        html = page.content()
        page.close()

        rows = safe_extract_prices(html)
        if rows:
            print(f"    ✅ Tìm thấy {len(rows)} data points")
        else:
            print(f"    ⚠️  Bị chặn hoặc không có chart data (HTML = {len(html)/1024:.1f}KB)")
            
        return rows
    except Exception as e:
        print(f"    ❌ Lỗi: {e}")
        try: page.close()
        except: pass
        return []

# ============================================================
# EXPORT
# ============================================================

def get_5y_months() -> list[str]:
    now = date.today()
    months = []
    for i in range(60, -1, -1):
        m = now.month - (i % 12)
        y = now.year - (i // 12)
        while m <= 0:
            m += 12
            y -= 1
        months.append(f"{y:04d}-{m:02d}")
    return sorted(list(set(months)))

def save_csv(pivot, months, path):
    import csv
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Tháng"] + list(pivot.keys()))
        for m in months:
            w.writerow([m] + [pivot[d].get(m, "") for d in pivot.keys()])
    print(f"✅ Đã lưu CSV: {path}")

def save_excel(pivot, months, path):
    if not HAS_EXCEL: return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Giá CC Theo Quận"
    
    from openpyxl.styles import PatternFill, Font, Alignment
    hf = PatternFill("solid", fgColor="1C3557")
    font = Font(color="FFFFFF", bold=True)
    al = Alignment(horizontal="center")
    
    quans = list(pivot.keys())
    
    ws.cell(1, 1, "Tháng").fill = hf
    ws.cell(1, 1).font = font
    
    for i, q in enumerate(quans, 2):
        ws.cell(1, i, q).fill = hf
        ws.cell(1, i).font = font
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 15
        
    for r, m in enumerate(months, 2):
        ws.cell(r, 1, m)
        for c, q in enumerate(quans, 2):
            ws.cell(r, c, pivot[q].get(m, None))
            
    ws.freeze_panes = "B2"
    wb.save(path)
    print(f"✅ Đã lưu Excel: {path}")

# ============================================================
# MAIN EXECUTOR
# ============================================================

def main():
    print("=" * 60)
    print("  CRAWL LỊCH SỬ GIÁ CHUNG CƯ 5 NĂM — batdongsan.com.vn")
    print("=" * 60)

    months_5y = get_5y_months()
    all_data = {q: {} for q in DISTRICTS.keys()}
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=["--disable-blink-features=AutomationControlled"])
        ctx = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36",
            viewport={"width": 1920, "height": 1080}
        )
        
        for name, slug in list(DISTRICTS.items())[:3]:  # Chỉ chạy thử 3 quận đầu
            rows = crawl_district_with_playwright(slug, name, ctx)
            for r in rows:
                if r["thang"] in months_5y:
                    all_data[name][r["thang"]] = r["gia_trm2"]
            time.sleep(random.uniform(3, 7))
            
        browser.close()

    if any(len(d)>0 for d in all_data.values()):
        save_csv(all_data, months_5y, OUTPUT_DIR / "lich_su_gia_cc.csv")
        save_excel(all_data, months_5y, OUTPUT_DIR / "lich_su_gia_cc.xlsx")
    else:
        print("❌ Không có dữ liệu để lưu do web chặn. Cần đổi Proxy hoặc IP.")

if __name__ == "__main__":
    main()
